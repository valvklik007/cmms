import re, os, random
import string
from dataclasses import dataclass
from datetime import timedelta
from re import template
from flask import request, render_template, url_for, session, redirect, flash, Response
import datetime
from dateutil.relativedelta import relativedelta

from sqlalchemy.inspection import inspect
from sqlalchemy import ForeignKey, and_, or_

from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from app import db, turbo, app
import inspect

from cmms import models



def request_db(data):
    for m in inspect.getmembers(models):
        if data.lower() == m[0].lower():
            return getattr(models, m[0])


def req_litter(req):
    request_key = {}
    for key, value in req.form.items():
        request_key.update({key: value})
    return request_key


class BaseRepository:
    def __init__(self, model):
        self._db = db.session
        self._model = model

    def add(self, **kwargs):
        try:
            item = self._model(**kwargs)
            self._db.add(item)
            self._db.commit()
            return item
        except Exception as e:
            self._db.rollback()
        return False

    def update(self, item, **kwargs):
        # item = self._model.query.filter_by(id=id_i)
        try:
            item.update({**kwargs})
            self._db.commit()
            return True
        except Exception as e:
            print(e)
            self._db.rollback()
        return False

    def delete(self, id_item):
        item = self._model.query.filter_by(id = id_item).first()
        if item:
            try:
                self._db.delete(item)
                self._db.commit()
                return True
            except Exception as e:
                self._db.rollback()
        return False

    def que(self):
        #запрос
        return self._db.query(self._model)

    def get_all(self):
        # Выбора всей таблицы
        return self._db.query(self._model).all()

    def get_filter(self, **kwargs):
        #Сформулировать запрос и вывести
        return self._db.query(self._model).filter_by(**kwargs)

    def property_parent(self, name):
        relationship = getattr(self._model, name)
        # Имя родителя
        parent_table_name = relationship.property.mapper.class_.__table__.name
        return parent_table_name

    # def get_attributes(self):
    #     attributes = []
    #     for column in self._model.__table__.columns:
    #         attr_info = {
    #             "name": column.name,
    #             "type": str(column.type),
    #             "is_foreign_key": is_foreign_key(column),  # Проверяем, является ли внешним ключом
    #         }
    #         attributes.append(attr_info)
    #     return attributes


class ManufacturerItme(BaseRepository):
    def __init__(self):
        super().__init__(models.ManufacturerItme)

    # def add_parent(self, name, fk):
    #     self.add(name=name, fk_ModelItme=fk)



class ModelItme(BaseRepository):
    def __init__(self):
        super().__init__(models.ModelItme)


class StatusItme(BaseRepository):
    def __init__(self):
        super().__init__(models.StatusItme)


class ConditionItme(BaseRepository):
    def __init__(self):
        super().__init__(models.ConditionItme)


class TypeEquipment(BaseRepository):
    model = models.TypeEquipment
    def __init__(self):
        super().__init__(self.model)


class PlaceOperation(BaseRepository):
    def __init__(self):
        super().__init__(models.PlaceOperation)

class Item(BaseRepository):
    model = models.Item
    def __init__(self):
        super().__init__(self.model)


class Notepad(BaseRepository):
    def __init__(self):
        super().__init__(models.Notepad)

class ManualBook(BaseRepository):
    def __init__(self):
        super().__init__(models.ManualBook)

class Schedules(BaseRepository):
    def __init__(self):
        super().__init__(models.Schedules)

class RegulatoryWork(BaseRepository):
    def __init__(self):
        super().__init__(models.RegulatoryWork)


class Maintenance(BaseRepository):
    model = models.Maintenance
    def __init__(self):
        super().__init__(self.model)

    def get_exist(self, id_item, id_maintenance):
        q = db.session.query(self.model).filter(
            self.model.id == id_maintenance,
            self.model.fk_item == id_item,
        ).scalar()
        return q


name_table ={
    'manufactureritme' : ManufacturerItme(),
    'modelitme' : ModelItme(),
    'statusitme' : StatusItme(),
    'conditionitme': ConditionItme(),
    'typeequipment' : TypeEquipment(),
    'placeoperation' : PlaceOperation(),
    'regulatorywork' : RegulatoryWork(),
    'schedules': Schedules(),
}


"""Контекст для карточки товара"""
def context_template_item():
    # Получаем контекст
    tem_context = dict(
        manufactureritme=ManufacturerItme().get_all(),
        modelitme=ModelItme().get_all(),
        statusitme=StatusItme().get_all(),
        conditionitme=ConditionItme().get_all(),
        typeequipment=TypeEquipment().get_all(),
        placeoperation=PlaceOperation().get_all(),
        data_time=datetime.datetime.now().strftime("%Y-%m-%d"),
    )
    return tem_context


"""Создания таблицы excel"""
def form_table(q, req):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Заголовки
    headers = ["Имя станка", "Тип станка", "Регламент работ" , "Наименования обслуживания" , "Частота ТО"]

    start_date = datetime.datetime.strptime(req.values.get('data_form'), '%Y-%m-%d')
    end_date = datetime.datetime.strptime(req.values.get('data_to'), '%Y-%m-%d')

    date_headers = [(start_date + timedelta(days=i)).strftime("%d-%b") for i in range((end_date - start_date).days + 1)]
    ws.append(headers)

    # Записываем заголовки дат по вертикали в столбец D (длина заголовка, плюс один)
    for i, date in enumerate(date_headers, start=len(headers)+1):
        cell = ws.cell(row=1, column=i, value=date)
        cell.alignment = Alignment(textRotation=90)


    max_lengths = {i: len(str(header)) for i, header in enumerate(headers, start=1)}

    #Фиксируем столбец для scrolling
    column_letter = get_column_letter(len(headers)+1)
    ws.freeze_panes = f"{column_letter}2"

    # Заполняем строки
    for record in q.all():
        row = [
            record.item.name,
            record.item.model_TypeEquipment.name,
            record.regulatory_work.name,
            record.name,
            record.schedules.name,
        ]

        #Ищем длину текста, что бы присвоить значения столбца
        for i, header in enumerate(row, start=1):
            if len(header) > max_lengths[i]:
                max_lengths[i] = len(header)
                print(max_lengths)


        # Генерируем пустые ячейки для дат
        date_cells = [""] * len(date_headers)

        # Логика установки "X" по частоте обслуживания
        task_start = record.data_time.strftime("%d-%b")

        for i, date_str in enumerate(date_headers):
            if task_start == date_str:
                date_cells[i] = "X"
            # else:
            #     date_cells[i] = "-"

        row.extend(date_cells)
        ws.append(row)


    # Автоматически настраиваем ширину столбцов
    for col_idx, max_length in max_lengths.items():
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max_length + 3

    # Автоматически настраиваем ширину столбцов для дат
    for i, len_head in enumerate(range(len(date_headers)), start=len(headers)+1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = 3

    # Сохраняем файл
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
    # wb.save("maintenance_schedule.xlsx")
    # print("Файл сохранен: maintenance_schedule.xlsx")



"""Экспорт в файл по выбранным значениям"""
def export():
    context = dict(
        typeequipment = TypeEquipment().get_all(),
    )
    if request.values.get('id'):
        q = Item.model.query.filter(
            Item.model.fk_TypeEquipment.in_(request.values.getlist('id'))
        ).all()
        context.update(
            item = q,
        )
    if request.values.get('item'):
        print(request.values.get('item'))
        q = Maintenance.model.query.filter(
            Maintenance.model.fk_item.in_(request.values.getlist('item')),
            Maintenance.model.data_time.between(request.values.get('data_form'), request.values.get('data_to'))
        ).order_by( Maintenance.model.fk_item.asc() )

        buffer = form_table(q=q,req=request)

        #Кодируем кириллицу, кодировка RFC 5987
        filename = quote(f"Отчет от {request.values.get('data_form')}-{request.values.get('data_to')}.xlsx")

        # Отправляем файл пользователю
        return Response(
            buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )

    # "Content-Disposition": "attachment; filename=Отчет от {}-{}.xlsx".format(request.values.get('data_form'), request.values.get('data_to')),
    return render_template('/cmms/export.html', **context)

"""Главная страница"""
def index():
    context = dict(
        item = Item().get_all(),
        maint = Maintenance.model.query.filter(
            Maintenance.model.check_work == False
        ).order_by(
            Maintenance.model.check_work.asc(),
            Maintenance.model.data_time.asc(),
        ).all(),
        new_data=datetime.datetime.now().date() + timedelta(days=7),
    )

    return render_template('/cmms/index.html', **context)


"Добавления карточки оборудования"
def add_item():
    tem_context = context_template_item()
    if request.method == 'POST':
        rq = req_litter(request)
        rq['exploitation'] = datetime.datetime.strptime(rq['exploitation'], '%Y-%m-%d').date()
        rq['warranty'] = datetime.datetime.strptime(rq['warranty'], '%Y-%m-%d').date()
        rq['date_product'] = datetime.datetime.strptime(rq['date_product'], '%Y-%m-%d').date()
        id_add = Item().add(**rq)
        if id_add:
            flash('Успешно добавили', 'info')
            return redirect(url_for('index.render_item_and_update_and_delete', id=id_add.id), 303)
        else:
            flash('Ошибка добавления', 'error')
            return redirect(url_for('index.add_item'))
    return render_template('/cmms/item.html', **tem_context)


"Редактирования и удаления карточки оборудования"
def render_item_and_update_and_delete(id):
    item = Item().que().get(id)
    if item is not None:
        if request.method == 'GET':
            tem_context = context_template_item()
            item = Item().get_filter(id=id).first()
            tem_context.update(item = item)
            return render_template('/cmms/item.html', **tem_context)
        if request.method == 'PUT':
            rq = req_litter(request)

            rq['exploitation'] = datetime.datetime.strptime(rq['exploitation'], '%Y-%m-%d').date()
            rq['warranty'] = datetime.datetime.strptime(rq['warranty'], '%Y-%m-%d').date()
            rq['date_product'] = datetime.datetime.strptime(rq['date_product'], '%Y-%m-%d').date()
            item = Item().get_filter(id=id)
            Item().update(item, **rq)
            flash('Успешно обновлены данные', 'info')
            return redirect(url_for('index.render_item_and_update_and_delete', id=id), 303)
        if request.method == 'DELETE':
            Item().delete(id)
            flash('Успешно удалено', 'info')
            return redirect(url_for('index.add_item'), 303)
    else:
        flash('Такой карточки оборудования нет', 'info')
        return redirect(url_for('index.add_item'))


"""Блокнот карточки"""
def notepad(id):
    item = Item().que().get(id)
    if item is not None:
        if request.method == 'POST':
            rq = req_litter(request)
            q = Notepad().get_filter(fk_item=id)
            if q.first() is not None:
                q = Notepad().get_filter(fk_item=id)
                Notepad().update(q, **rq)
            else:
                Notepad().add(**rq)
        return render_template('/cmms/notepad.html', item=item)
    else:
        return redirect(url_for('index.add_item'))


"""Меню обслуживания"""
def maintenance_menu(id):
    item = Item().que().get(id)

    if request.values.get('delete') is not None:
        Maintenance().delete(request.values.get('delete'))
        return redirect(url_for('index.maintenance_menu', id=id), 303)

    if request.values.get('work') is not None:
        # переводим статус в выполнено check_work
        i = Maintenance().get_filter(id=request.values.get('work'))

        #Если выполнили раньше, делаем update даты, что бы выполнено не было в будущем
        if datetime.datetime.now().date() < i.first().data_time:
            Maintenance().update(i, data_time=datetime.datetime.now().date())

        Maintenance().update(i, check_work=True)

        delta = deltatime(i.first().schedules.period, +i.first().schedules.date_time)
        con = dict(
            name= i.first().name,
            data_time = datetime.datetime.now().date() + delta,
            data_time_old = i.first().data_time,
            fk_schedule = i.first().fk_schedule,
            fk_regulatory_work = i.first().fk_regulatory_work,
            fk_item = i.first().fk_item,
        )
        Maintenance().add(**con)

        return redirect(url_for('index.maintenance_menu', id=id), 303)

    if item is not None:
        context = dict(
            q = Maintenance.model.query.filter(
                Maintenance.model.fk_item == id,
            ).order_by(
                Maintenance.model.check_work.asc(),
                Maintenance.model.data_time.asc(),
            ),
            new_data =  datetime.datetime.now().date() + timedelta(days=7),
            item=item,
        )
    else:
        return redirect(url_for('index.add_item'))
    return render_template('/cmms/service.html', **context)


"""Даты"""
def deltatime(i, number):
    deltatime = {
        'days': relativedelta(days=number),
        'weeks': relativedelta(weeks=number),
        'months': relativedelta(months=number),
        'years': relativedelta(years=number),
    }
    return deltatime[i]


"""Обслуживание"""
def maintenance(id, id_maintenance):
    #Если id_maintenance == 0, то это является созданием обслуживания
    item = Item().que().get(id)
    maintenance_c = Maintenance()
    #Проверяем, есть ли указная карточка, либо делаем redirect на добавлении карточки
    if item is not None:
        context = dict(
            data_time = datetime.datetime.now().strftime("%Y-%m-%d"),
            maintenance = maintenance_c.que().get(id_maintenance),
            schedules = Schedules().get_all(),
            regulatorywork = RegulatoryWork().get_all(),
            item = item,
            url = request.url
        )
        if request.method != 'GET':
            rq = req_litter(request)
            check_work = datetime.datetime.strptime(rq['data_time'], '%Y-%m-%d').date()
            q = Schedules().get_filter(id=rq['fk_schedule']).first()
            rq['data_time'] = check_work + deltatime(q.period, +q.date_time)
            rq['fk_item'] = id
            rq.update(data_time_old = check_work)

        # id_maintenance == 0, 0 это создания обслуживания
        if id_maintenance is 0:
            if request.method == 'POST':
                add = maintenance_c.add(**rq)
                flash('Добавлено', 'info')
                return redirect(url_for('index.maintenance', id=id, id_maintenance=add.id), 303)
        else:
            # Проверяем, принадлежит ли запись к данной карточки, если нет, то делаем redirect на id_maintenance = 0
            q = maintenance_c.get_exist(id_item=id, id_maintenance=id_maintenance)
            if q:
                if request.method == 'PUT':
                    print(request.form)
                    main = maintenance_c.get_filter(id=id_maintenance)
                    maintenance_c.update(main, **rq)
                    flash('Отредактировано', 'info')
                    return redirect(url_for('index.maintenance', id=id, id_maintenance=id_maintenance), 303)
            else:
                return redirect(url_for('index.maintenance', id=id, id_maintenance=0), 303)

        return render_template('cmms/to_add_edit.html', **context)
    return redirect(url_for('index.add_item'))


"""Рандом"""
def ran(i):
    a = string.ascii_lowercase
    c = string.digits
    n = ""
    for sim_1 in range(i):
        n = n + random.choice(a + c)
    return n


"""Документация"""
def manual_book(id):
    item = Item().que().get(id)
    if item is not None:
        if request.method == 'DELETE':
            print(request.form['delete'])
            ManualBook().delete(request.form['delete'])
            return redirect(url_for('index.manual_book', id=item.id), 201)
        elif request.method == 'POST':
            file = request.files.getlist('files')
            for m in file:
                re_f = re.sub(r'[\]\[\/\\;,><&*:%=+@!#{}^()|?^ ]', '-', m.filename)
                file = f"{item.name}-{ran(10)}-{re_f}"
                m.save(os.path.join(app.config['UPLOAD_FOLDER'], file))
                ManualBook().add(**dict(slug = file, fk_item = id,))

            return redirect(url_for('index.manual_book', id=item.id))

        return render_template('/cmms/manual.html', item=item)
    return redirect(url_for('index.add_item'))


"Стримы turbo flask"
def push_update(template, **kwargs):
    turbo.push(
        turbo.update(
            render_template(
                template,
                **kwargs),
            target=kwargs['target'],),
    )

def push_append(template, **kwargs):
    turbo.push(
        turbo.append(
            render_template(
                template,
                **kwargs),
            target=kwargs['target'],),
    )


"Общая модификация параметров. Добавления, удаления, однотипных таблиц"
def add_delete_option(data):
    parent_records = None
    data_dict = dict(
        target=data,
    )
    if request.method == "POST":
        if request.form['action'] == 'add':
            request_key = {}

            date_time = {
                "days": "days",
                "weeks": "weeks",
                "months": "months",
                "years": "years",
            }

            if data == 'schedules':
                if not request.form.get('period'):
                    return []
                elif not date_time.get(request.form['period']):
                    return []

            for key, value in request.form.items():
                if 'action' not in key:
                    request_key.update({key: value})

            it = name_table[data].add(**request_key)
            if not it:
                flash('Ошибка', 'error')
                return redirect(url_for('index.add_delete_option', data=data), 200)
            data_dict.update(option_values = name_table[data].get_filter(**request_key).all(),)

            if turbo.can_stream():
                push_append(template='/cmms/include/add_option_value.html', **data_dict),
                return []

        elif request.form['action'] == 'delete':
            for m in request.form.getlist('allocated_name'):
                name_table[data].delete(m)

    if request.values.get('id') is not None:
        if 'modelitme' in data:
            option_values = name_table[data].get_filter(fk_ModelItme = request.values.get('id')).all()
            parent_records = name_table[name_table[data].property_parent('parent')].get_all()
    else:
        option_values = name_table[data].get_all()
        print(option_values)

    data_dict.update(
        option_values=option_values,
        parent_records=parent_records,
        id_parent = request.values.get('id'),
        url = data,
    )

    if turbo.can_stream():
        push_update(template='/cmms/include/add_option_value.html', **data_dict),
        return []
    else:
        return render_template('/cmms/add_option.html', **data_dict)

